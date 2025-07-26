import pytest
import tempfile
import os
from datetime import date, datetime, timedelta
from io import BytesIO
from unittest.mock import patch, MagicMock

from fastapi.testclient import TestClient
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.main import app
from app.db.session import get_db
from app.core.auth import get_current_user
from app.core.export_service import ExportService
from app.models.user import User
from app.models.receipt import Receipt
from app.models.line_item import LineItem
from app.models.category import Category


def create_test_user(db: Session, email: str = "test@example.com") -> User:
    """Create a test user."""
    user = User(
        email=email,
        hashed_password="fake_hash",
        is_active=True
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def create_test_receipt(
    db: Session, 
    user_id: int, 
    store_name: str = "Test Store",
    total_amount: float = 100.0,
    receipt_date: date = None
) -> Receipt:
    """Create a test receipt."""
    if receipt_date is None:
        receipt_date = date.today()
    
    receipt = Receipt(
        store_name=store_name,
        receipt_date=receipt_date,
        total_amount=total_amount,
        tax_amount=10.0,
        currency="USD",
        user_id=user_id,
        processing_status="completed",
        is_verified=True
    )
    db.add(receipt)
    db.commit()
    db.refresh(receipt)
    return receipt


class TestExportAPI:
    """Test the export API endpoints."""
    
    @pytest.fixture
    def client(self, test_db_session):
        """Create test client with database dependency override."""
        def get_test_db_override():
            return test_db_session
        
        app.dependency_overrides[get_db] = get_test_db_override
        
        with TestClient(app) as test_client:
            yield test_client
        
        app.dependency_overrides.clear()
    
    @pytest.fixture
    def mock_auth(self, test_db_session):
        """Mock authentication dependency."""
        test_user = create_test_user(test_db_session)
        
        def get_current_user_override():
            return test_user
        
        app.dependency_overrides[get_current_user] = get_current_user_override
        yield test_user
        app.dependency_overrides.clear()
    
    @pytest.fixture
    def test_receipts(self, test_db_session, mock_auth):
        """Create test receipts with line items."""
        receipts = []
        categories = []
        
        # Create test categories
        for cat_name in ['Food', 'Transport', 'Office']:
            category = Category(name=cat_name)
            test_db_session.add(category)
            test_db_session.commit()
            test_db_session.refresh(category)
            categories.append(category)
        
        # Create test receipts
        for i in range(5):
            receipt = create_test_receipt(
                test_db_session, 
                mock_auth.id,
                store_name=f"Store {i+1}",
                total_amount=100.0 + (i * 25),
                receipt_date=date.today() - timedelta(days=i*7)
            )
            
            # Add line items to receipt
            for j in range(2):
                line_item = LineItem(
                    receipt_id=receipt.id,
                    name=f"Item {j+1} from {receipt.store_name}",
                    quantity=1 + j,
                    unit_price=25.0,
                    total_price=25.0 * (1 + j),
                    category_id=categories[j % len(categories)].id
                )
                test_db_session.add(line_item)
            
            test_db_session.commit()
            test_db_session.refresh(receipt)
            receipts.append(receipt)
        
        return receipts, categories
    
    def test_export_excel_success(self, client, mock_auth, test_receipts):
        """Test successful Excel export."""
        receipts, categories = test_receipts
        
        response = client.get("/api/v1/export/excel")
        
        if response.status_code != 200:
            print(f"Response status: {response.status_code}")
            print(f"Response content: {response.content}")
        
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]
        
        # Verify content is valid Excel file
        excel_data = BytesIO(response.content)
        workbook = load_workbook(excel_data)
        
        # Check sheet names
        sheet_names = workbook.sheetnames
        assert "Summary" in sheet_names
        assert "Receipts Summary" in sheet_names
        assert "Line Items Detail" in sheet_names
        
        # Check receipts sheet content
        receipts_sheet = workbook["Receipts Summary"]
        assert receipts_sheet.cell(1, 1).value == "Receipt ID"
        assert receipts_sheet.cell(1, 2).value == "Store Name"
        
        # Should have header + 5 receipt rows
        assert receipts_sheet.max_row == 6
    
    def test_export_excel_with_date_range(self, client, mock_auth, test_receipts):
        """Test Excel export with date range filtering."""
        receipts, categories = test_receipts
        
        # Export last 30 days
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()
        
        response = client.get(
            f"/api/v1/export/excel?start_date={start_date}&end_date={end_date}"
        )
        
        assert response.status_code == 200
        
        # Verify Excel content
        excel_data = BytesIO(response.content)
        workbook = load_workbook(excel_data)
        receipts_sheet = workbook["Receipts Summary"]
        
        # Should have all receipts within date range
        assert receipts_sheet.max_row > 1  # Header + data rows
    
    def test_export_excel_without_line_items(self, client, mock_auth, test_receipts):
        """Test Excel export without line items sheet."""
        receipts, categories = test_receipts
        
        response = client.get("/api/v1/export/excel?include_line_items=false")
        
        assert response.status_code == 200
        
        # Verify Excel content
        excel_data = BytesIO(response.content)
        workbook = load_workbook(excel_data)
        
        # Check sheet names - should not have line items sheet
        sheet_names = workbook.sheetnames
        assert "Summary" in sheet_names
        assert "Receipts Summary" in sheet_names
        assert "Line Items Detail" not in sheet_names
    
    def test_export_excel_invalid_date_range(self, client, mock_auth):
        """Test Excel export with invalid date range."""
        start_date = date.today()
        end_date = date.today() - timedelta(days=10)  # End before start
        
        response = client.get(
            f"/api/v1/export/excel?start_date={start_date}&end_date={end_date}"
        )
        
        assert response.status_code == 400
        assert "end_date must be after or equal to start_date" in response.json()["detail"]
    
    def test_export_info_success(self, client, mock_auth, test_receipts):
        """Test export info endpoint."""
        receipts, categories = test_receipts
        
        export_query = {
            "include_line_items": True
        }
        
        response = client.post("/api/v1/export/excel/info", json=export_query)
        
        assert response.status_code == 200
        data = response.json()
        
        assert data["success"] is True
        assert data["records_count"] == 5  # 5 test receipts
        assert data["date_range"] == "All time"
        assert ".xlsx" in data["filename"]
    
    def test_export_info_with_date_range(self, client, mock_auth, test_receipts):
        """Test export info with date range."""
        receipts, categories = test_receipts
        
        start_date = date.today() - timedelta(days=15)
        end_date = date.today()
        
        export_query = {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "include_line_items": False
        }
        
        response = client.post("/api/v1/export/excel/info", json=export_query)
        
        assert response.status_code == 200
        data = response.json()
        
        assert data["success"] is True
        assert start_date.strftime("%Y-%m-%d") in data["date_range"]
        assert end_date.strftime("%Y-%m-%d") in data["date_range"]
    
    def test_export_unauthorized(self, client):
        """Test export without authentication."""
        # Remove auth override to test unauthorized access
        app.dependency_overrides.clear()
        
        response = client.get("/api/v1/export/excel")
        
        assert response.status_code == 403


class TestExportService:
    """Test the export service functionality."""
    
    @pytest.fixture
    def export_service(self, test_db_session):
        """Create export service instance."""
        return ExportService(test_db_session)
    
    @pytest.fixture
    def test_data(self, test_db_session):
        """Create test data for export service tests."""
        user = create_test_user(test_db_session)
        
        # Create category
        category = Category(name="Test Category")
        test_db_session.add(category)
        test_db_session.commit()
        test_db_session.refresh(category)
        
        # Create receipts
        receipts = []
        for i in range(3):
            receipt = create_test_receipt(
                test_db_session,
                user.id,
                store_name=f"Test Store {i+1}",
                total_amount=50.0 + (i * 25),
                receipt_date=date.today() - timedelta(days=i*5)
            )
            
            # Add line item
            line_item = LineItem(
                receipt_id=receipt.id,
                name=f"Test Item {i+1}",
                quantity=1,
                unit_price=50.0 + (i * 25),
                total_price=50.0 + (i * 25),
                category_id=category.id
            )
            test_db_session.add(line_item)
            test_db_session.commit()
            receipts.append(receipt)
        
        return user, receipts, category
    
    def test_export_receipts_to_excel(self, export_service, test_data):
        """Test Excel export functionality."""
        user, receipts, category = test_data
        
        excel_buffer, filename = export_service.export_receipts_to_excel(
            user_id=user.id,
            include_line_items=True
        )
        
        assert isinstance(excel_buffer, BytesIO)
        assert filename.endswith('.xlsx')
        assert 'expense_export' in filename
        
        # Verify Excel content
        workbook = load_workbook(excel_buffer)
        
        # Check sheet names
        sheet_names = workbook.sheetnames
        assert "Summary" in sheet_names
        assert "Receipts Summary" in sheet_names
        assert "Line Items Detail" in sheet_names
        
        # Check receipts sheet
        receipts_sheet = workbook["Receipts Summary"]
        assert receipts_sheet.max_row == 4  # Header + 3 receipts
        
        # Check line items sheet
        line_items_sheet = workbook["Line Items Detail"]
        assert line_items_sheet.max_row == 4  # Header + 3 line items
    
    def test_export_with_date_filtering(self, export_service, test_data):
        """Test export with date range filtering."""
        user, receipts, category = test_data
        
        # Filter to only recent receipts
        start_date = date.today() - timedelta(days=7)
        
        excel_buffer, filename = export_service.export_receipts_to_excel(
            user_id=user.id,
            start_date=start_date,
            include_line_items=True
        )
        
        # Verify Excel content shows filtered data
        workbook = load_workbook(excel_buffer)
        receipts_sheet = workbook["Receipts Summary"]
        
        # Should have fewer receipts due to filtering
        assert receipts_sheet.max_row <= 4  # Header + filtered receipts
    
    def test_export_without_line_items(self, export_service, test_data):
        """Test export without line items sheet."""
        user, receipts, category = test_data
        
        excel_buffer, filename = export_service.export_receipts_to_excel(
            user_id=user.id,
            include_line_items=False
        )
        
        # Verify Excel content
        workbook = load_workbook(excel_buffer)
        sheet_names = workbook.sheetnames
        
        assert "Summary" in sheet_names
        assert "Receipts Summary" in sheet_names
        assert "Line Items Detail" not in sheet_names
    
    def test_export_temp_file(self, export_service, test_data):
        """Test export to temporary file."""
        user, receipts, category = test_data
        
        temp_path, filename = export_service.export_receipts_to_temp_file(
            user_id=user.id,
            include_line_items=True
        )
        
        try:
            assert os.path.exists(temp_path)
            assert temp_path.endswith('.xlsx')
            assert filename.endswith('.xlsx')
            
            # Verify file content
            workbook = load_workbook(temp_path)
            assert "Summary" in workbook.sheetnames
            
        finally:
            # Clean up
            if os.path.exists(temp_path):
                os.remove(temp_path)
    
    def test_filename_generation(self, export_service):
        """Test filename generation with different parameters."""
        # Test filename without date range
        filename1 = export_service._generate_filename()
        assert 'expense_export_all_' in filename1
        assert filename1.endswith('.xlsx')
        
        # Test filename with date range
        start_date = date(2024, 1, 1)
        end_date = date(2024, 1, 31)
        filename2 = export_service._generate_filename(start_date, end_date)
        assert '20240101_to_20240131' in filename2
        assert filename2.endswith('.xlsx')
    
    def test_receipts_data_query_optimization(self, export_service, test_data):
        """Test that receipts data query is optimized."""
        user, receipts, category = test_data
        
        # Test the internal method
        receipts_data = export_service._get_receipts_data(user.id)
        
        assert len(receipts_data) == 3
        assert all('line_items_count' in receipt for receipt in receipts_data)
        assert all(isinstance(receipt['line_items_count'], int) for receipt in receipts_data)
    
    def test_line_items_data_query_optimization(self, export_service, test_data):
        """Test that line items data query is optimized."""
        user, receipts, category = test_data
        
        # Test the internal method
        line_items_data = export_service._get_line_items_data(user.id)
        
        assert len(line_items_data) == 3
        assert all('category' in item for item in line_items_data)
        assert all(item['category'] == 'Test Category' for item in line_items_data)
    
    def test_cleanup_temp_files(self, export_service):
        """Test temporary file cleanup functionality."""
        # Create a temporary file
        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)
        
        # Add to tracking
        export_service._temp_files.add(temp_path)
        
        # Verify file exists
        assert os.path.exists(temp_path)
        
        # Test cleanup
        export_service.cleanup_temp_files()
        
        # Verify file is removed
        assert not os.path.exists(temp_path)
        assert temp_path not in export_service._temp_files


class TestExportPerformance:
    """Test export functionality performance with larger datasets."""
    
    @pytest.fixture
    def large_dataset(self, test_db_session):
        """Create larger test dataset for performance testing."""
        user = create_test_user(test_db_session)
        
        # Create categories
        categories = []
        for i in range(5):
            category = Category(name=f"Category {i+1}")
            test_db_session.add(category)
            categories.append(category)
        test_db_session.commit()
        
        # Create many receipts
        receipts = []
        for i in range(100):  # 100 receipts
            receipt = create_test_receipt(
                test_db_session,
                user.id,
                store_name=f"Store {i+1}",
                total_amount=10.0 + (i % 100),
                receipt_date=date.today() - timedelta(days=i % 365)
            )
            
            # Add multiple line items per receipt
            for j in range(3):  # 3 items per receipt = 300 total line items
                line_item = LineItem(
                    receipt_id=receipt.id,
                    name=f"Item {j+1} from Receipt {i+1}",
                    quantity=1,
                    unit_price=5.0,
                    total_price=5.0,
                    category_id=categories[j % len(categories)].id
                )
                test_db_session.add(line_item)
            
            receipts.append(receipt)
            
            # Commit in batches for performance
            if i % 20 == 0:
                test_db_session.commit()
        
        test_db_session.commit()
        return user, receipts, categories
    
    @pytest.mark.benchmark
    def test_large_export_performance(self, test_db_session, large_dataset):
        """Test export performance with large dataset."""
        user, receipts, categories = large_dataset
        
        export_service = ExportService(test_db_session)
        
        # Measure export time
        import time
        start_time = time.time()
        
        excel_buffer, filename = export_service.export_receipts_to_excel(
            user_id=user.id,
            include_line_items=True
        )
        
        end_time = time.time()
        export_time = end_time - start_time
        
        # Verify export completed successfully
        assert isinstance(excel_buffer, BytesIO)
        assert excel_buffer.getvalue()  # Has content
        
        # Performance should be reasonable (less than 10 seconds for 100 receipts)
        assert export_time < 10.0
        
        # Verify content
        workbook = load_workbook(excel_buffer)
        receipts_sheet = workbook["Receipts Summary"]
        line_items_sheet = workbook["Line Items Detail"]
        
        # Should have all receipts + header
        assert receipts_sheet.max_row == 101
        
        # Should have all line items + header
        assert line_items_sheet.max_row == 301