from datetime import datetime, date
from typing import List, Optional, Tuple, BinaryIO
from io import BytesIO
import tempfile
import os
import logging
from pathlib import Path
import time
import threading

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session
from sqlalchemy import and_, func

from app.models.receipt import Receipt
from app.models.line_item import LineItem
from app.models.category import Category
from app.models.user import User

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting receipt and expense data to Excel format."""
    
    def __init__(self, db: Session):
        self.db = db
        self._temp_files = set()  # Track temporary files for cleanup
        
    def cleanup_temp_files(self):
        """Clean up any temporary files created during export."""
        for temp_file in list(self._temp_files):
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    logger.debug(f"Cleaned up temporary file: {temp_file}")
                self._temp_files.discard(temp_file)
            except Exception as e:
                logger.warning(f"Failed to clean up temporary file {temp_file}: {str(e)}")
    
    def schedule_cleanup(self, delay_seconds: int = 300):
        """Schedule cleanup of temporary files after a delay (default 5 minutes)."""
        def delayed_cleanup():
            time.sleep(delay_seconds)
            self.cleanup_temp_files()
        
        cleanup_thread = threading.Thread(target=delayed_cleanup, daemon=True)
        cleanup_thread.start()
        
    def export_receipts_to_excel(
        self, 
        user_id: int, 
        start_date: Optional[date] = None, 
        end_date: Optional[date] = None,
        include_line_items: bool = True
    ) -> Tuple[BytesIO, str]:
        """
        Export receipts to Excel format with optional date range filtering.
        
        Args:
            user_id: ID of the user whose receipts to export
            start_date: Optional start date for filtering
            end_date: Optional end date for filtering  
            include_line_items: Whether to include line items in a separate sheet
            
        Returns:
            Tuple of (BytesIO containing Excel data, filename)
        """
        try:
            # Get receipts data
            logger.debug(f"Getting receipts data for user {user_id}")
            receipts_data = self._get_receipts_data(user_id, start_date, end_date)
            
            # Create workbook
            logger.debug("Creating workbook")
            workbook = Workbook()
            
            # Create receipts sheet
            logger.debug("Creating receipts sheet")
            self._create_receipts_sheet(workbook, receipts_data)
            
            # Create line items sheet if requested
            if include_line_items:
                logger.debug("Getting line items data")
                line_items_data = self._get_line_items_data(user_id, start_date, end_date)
                logger.debug("Creating line items sheet")
                self._create_line_items_sheet(workbook, line_items_data)
            
            # Create summary sheet
            logger.debug("Creating summary sheet")
            self._create_summary_sheet(workbook, receipts_data, start_date, end_date)
            
            # Save to BytesIO
            logger.debug("Saving workbook to buffer")
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Generate filename
            filename = self._generate_filename(start_date, end_date)
            
            logger.info(f"Successfully exported {len(receipts_data)} receipts for user {user_id}")
            
            return excel_buffer, filename
            
        except Exception as e:
            logger.error(f"Error exporting receipts to Excel for user {user_id}: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise
    
    def export_receipts_to_temp_file(
        self, 
        user_id: int, 
        start_date: Optional[date] = None, 
        end_date: Optional[date] = None,
        include_line_items: bool = True
    ) -> Tuple[str, str]:
        """
        Export receipts to a temporary Excel file (useful for large datasets).
        
        Args:
            user_id: ID of the user whose receipts to export
            start_date: Optional start date for filtering
            end_date: Optional end date for filtering  
            include_line_items: Whether to include line items in a separate sheet
            
        Returns:
            Tuple of (temporary file path, filename)
        """
        try:
            # Get receipts data
            receipts_data = self._get_receipts_data(user_id, start_date, end_date)
            
            # Create workbook
            workbook = Workbook()
            
            # Create receipts sheet
            self._create_receipts_sheet(workbook, receipts_data)
            
            # Create line items sheet if requested
            if include_line_items:
                line_items_data = self._get_line_items_data(user_id, start_date, end_date)
                self._create_line_items_sheet(workbook, line_items_data)
            
            # Create summary sheet
            self._create_summary_sheet(workbook, receipts_data, start_date, end_date)
            
            # Generate filename
            filename = self._generate_filename(start_date, end_date)
            
            # Create temporary file
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx', prefix='export_')
            os.close(temp_fd)  # Close the file descriptor
            
            # Save workbook to temporary file
            workbook.save(temp_path)
            
            # Track temporary file for cleanup
            self._temp_files.add(temp_path)
            
            # Schedule cleanup
            self.schedule_cleanup()
            
            logger.info(f"Successfully exported {len(receipts_data)} receipts to temporary file for user {user_id}")
            
            return temp_path, filename
            
        except Exception as e:
            logger.error(f"Error exporting receipts to temporary file for user {user_id}: {str(e)}")
            raise
    
    def _get_receipts_data(
        self, 
        user_id: int, 
        start_date: Optional[date] = None, 
        end_date: Optional[date] = None
    ) -> List[dict]:
        """
        Get receipts data with optional date filtering and optimized queries.
        Uses subquery to efficiently count line items without loading all relationships.
        """
        
        # Subquery to count line items efficiently
        line_items_count_subquery = (
            self.db.query(func.count(LineItem.id))
            .filter(LineItem.receipt_id == Receipt.id)
            .scalar_subquery()
        )
        
        # Main query with optimized line item counting
        query = (
            self.db.query(
                Receipt.id,
                Receipt.store_name,
                Receipt.receipt_date,
                Receipt.total_amount,
                Receipt.currency,
                Receipt.processing_status,
                Receipt.is_verified,
                Receipt.created_at,
                line_items_count_subquery.label('line_items_count')
            )
            .filter(Receipt.user_id == user_id)
        )
        
        # Apply date filters
        if start_date:
            query = query.filter(Receipt.receipt_date >= start_date)
        if end_date:
            query = query.filter(Receipt.receipt_date <= end_date)
            
        # Order by receipt date for consistent export ordering
        receipts = query.order_by(Receipt.receipt_date.desc()).all()
        
        # Convert to dict format for Excel generation
        receipts_data = []
        for receipt in receipts:
            # Convert timezone-aware datetime to naive for Excel compatibility
            created_at = receipt.created_at
            if hasattr(created_at, 'tzinfo') and created_at.tzinfo is not None:
                created_at = created_at.replace(tzinfo=None)
                
            # Convert receipt_date to naive datetime if it's timezone-aware
            receipt_date = receipt.receipt_date
            if hasattr(receipt_date, 'tzinfo') and receipt_date.tzinfo is not None:
                receipt_date = receipt_date.replace(tzinfo=None)
                
            receipts_data.append({
                'id': receipt.id,
                'store_name': receipt.store_name,
                'receipt_date': receipt_date,
                'total_amount': float(receipt.total_amount),
                'currency': receipt.currency,
                'processing_status': receipt.processing_status,
                'is_verified': receipt.is_verified,
                'created_at': created_at,
                'line_items_count': receipt.line_items_count or 0
            })
            
        logger.debug(f"Retrieved {len(receipts_data)} receipts for user {user_id} with date range {start_date} to {end_date}")
        return receipts_data
    
    def _get_line_items_data(
        self, 
        user_id: int, 
        start_date: Optional[date] = None, 
        end_date: Optional[date] = None
    ) -> List[dict]:
        """
        Get line items data with optional date filtering and optimized joins.
        Uses efficient joins to minimize database queries and improve performance.
        """
        
        # Optimized query using explicit column selection to reduce data transfer
        query = (
            self.db.query(
                LineItem.id,
                LineItem.name.label('item_name'),
                LineItem.quantity,
                LineItem.unit_price,
                LineItem.total_price,
                LineItem.created_at,
                Receipt.id.label('receipt_id'),
                Receipt.store_name,
                Receipt.receipt_date,
                Category.name.label('category_name')
            )
            .join(Receipt, LineItem.receipt_id == Receipt.id)
            .outerjoin(Category, LineItem.category_id == Category.id)
            .filter(Receipt.user_id == user_id)
        )
        
        # Apply date filters on receipt date
        if start_date:
            query = query.filter(Receipt.receipt_date >= start_date)
        if end_date:
            query = query.filter(Receipt.receipt_date <= end_date)
            
        # Order by receipt date and line item for consistent export
        results = query.order_by(Receipt.receipt_date.desc(), LineItem.id).all()
        
        # Convert to dict format for Excel generation
        line_items_data = []
        for item in results:
            # Convert timezone-aware datetime to naive for Excel compatibility
            created_at = item.created_at
            if hasattr(created_at, 'tzinfo') and created_at.tzinfo is not None:
                created_at = created_at.replace(tzinfo=None)
                
            # Convert receipt_date to naive datetime if it's timezone-aware
            receipt_date = item.receipt_date
            if hasattr(receipt_date, 'tzinfo') and receipt_date.tzinfo is not None:
                receipt_date = receipt_date.replace(tzinfo=None)
                
            line_items_data.append({
                'receipt_id': item.receipt_id,
                'store_name': item.store_name,
                'receipt_date': receipt_date,
                'item_name': item.item_name,
                'quantity': float(item.quantity) if item.quantity else 1.0,
                'unit_price': float(item.unit_price) if item.unit_price else 0.0,
                'total_price': float(item.total_price),
                'category': item.category_name if item.category_name else 'Uncategorized',
                'created_at': created_at
            })
            
        logger.debug(f"Retrieved {len(line_items_data)} line items for user {user_id} with date range {start_date} to {end_date}")
        return line_items_data
    
    def _create_receipts_sheet(self, workbook: Workbook, receipts_data: List[dict]):
        """Create the receipts summary sheet."""
        
        # Remove default sheet and create receipts sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        ws = workbook.create_sheet('Receipts Summary')
        
        # Define headers
        headers = [
            'Receipt ID', 'Store Name', 'Receipt Date', 'Total Amount', 
            'Currency', 'Status', 'Verified', 'Items Count', 'Upload Date'
        ]
        
        # Apply header styling
        self._apply_header_style(ws, headers)
        
        # Add data rows
        for row_idx, receipt in enumerate(receipts_data, start=2):
            ws.cell(row=row_idx, column=1, value=receipt['id'])
            ws.cell(row=row_idx, column=2, value=receipt['store_name'])
            ws.cell(row=row_idx, column=3, value=receipt['receipt_date'])
            ws.cell(row=row_idx, column=4, value=receipt['total_amount'])
            ws.cell(row=row_idx, column=5, value=receipt['currency'])
            ws.cell(row=row_idx, column=6, value=receipt['processing_status'])
            ws.cell(row=row_idx, column=7, value='Yes' if receipt['is_verified'] else 'No')
            ws.cell(row=row_idx, column=8, value=receipt['line_items_count'])
            
            ws.cell(row=row_idx, column=9, value=receipt['created_at'])
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Apply data formatting
        self._apply_data_formatting(ws, len(receipts_data))
    
    def _create_line_items_sheet(self, workbook: Workbook, line_items_data: List[dict]):
        """Create the line items detail sheet."""
        
        ws = workbook.create_sheet('Line Items Detail')
        
        # Define headers
        headers = [
            'Receipt ID', 'Store Name', 'Receipt Date', 'Item Name', 
            'Quantity', 'Unit Price', 'Total Price', 'Category', 'Date Added'
        ]
        
        # Apply header styling
        self._apply_header_style(ws, headers)
        
        # Add data rows
        for row_idx, item in enumerate(line_items_data, start=2):
            ws.cell(row=row_idx, column=1, value=item['receipt_id'])
            ws.cell(row=row_idx, column=2, value=item['store_name'])
            ws.cell(row=row_idx, column=3, value=item['receipt_date'])
            ws.cell(row=row_idx, column=4, value=item['item_name'])
            ws.cell(row=row_idx, column=5, value=item['quantity'])
            ws.cell(row=row_idx, column=6, value=item['unit_price'])
            ws.cell(row=row_idx, column=7, value=item['total_price'])
            ws.cell(row=row_idx, column=8, value=item['category'])
            
            ws.cell(row=row_idx, column=9, value=item['created_at'])
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Apply data formatting
        self._apply_data_formatting(ws, len(line_items_data))
    
    def _create_summary_sheet(
        self, 
        workbook: Workbook, 
        receipts_data: List[dict], 
        start_date: Optional[date], 
        end_date: Optional[date]
    ):
        """Create a summary statistics sheet."""
        
        ws = workbook.create_sheet('Summary', 0)  # Insert at beginning
        
        # Calculate summary statistics
        total_receipts = len(receipts_data)
        total_amount = sum(receipt['total_amount'] for receipt in receipts_data)
        avg_amount = total_amount / total_receipts if total_receipts > 0 else 0
        verified_receipts = sum(1 for receipt in receipts_data if receipt['is_verified'])
        
        # Get date range
        date_range = "All time"
        if start_date or end_date:
            start_str = start_date.strftime("%Y-%m-%d") if start_date else "Beginning"
            end_str = end_date.strftime("%Y-%m-%d") if end_date else "Present"
            date_range = f"{start_str} to {end_str}"
        
        # Add summary data - ensure timezone-naive datetime for Excel compatibility
        export_datetime = datetime.now()
        if hasattr(export_datetime, 'tzinfo') and export_datetime.tzinfo is not None:
            export_datetime = export_datetime.replace(tzinfo=None)
        
        summary_data = [
            ['Export Summary', ''],
            ['Date Range:', date_range],
            ['Export Date:', export_datetime.strftime("%Y-%m-%d %H:%M:%S")],
            ['', ''],
            ['Statistics', ''],
            ['Total Receipts:', total_receipts],
            ['Total Amount:', f"${total_amount:.2f}"],
            ['Average Amount:', f"${avg_amount:.2f}"],
            ['Verified Receipts:', f"{verified_receipts} ({verified_receipts/total_receipts*100:.1f}%)" if total_receipts > 0 else "0 (0%)"],
        ]
        
        # Apply summary styling
        for row_idx, (label, value) in enumerate(summary_data, start=1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)
            
            # Style headers
            if label in ['Export Summary', 'Statistics']:
                cell = ws.cell(row=row_idx, column=1)
                cell.font = Font(bold=True, size=14)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=14)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _apply_header_style(self, ws: Worksheet, headers: List[str]):
        """Apply styling to header row."""
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
    
    def _apply_data_formatting(self, ws: Worksheet, data_rows: int):
        """Apply formatting to data rows."""
        
        if data_rows == 0:
            return
            
        # Apply borders and alternating row colors
        for row_idx in range(2, data_rows + 2):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                
                # Alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    def _auto_adjust_columns(self, ws: Worksheet):
        """Auto-adjust column widths based on content."""
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _generate_filename(
        self, 
        start_date: Optional[date] = None, 
        end_date: Optional[date] = None
    ) -> str:
        """Generate a meaningful filename for the export."""
        
        # Ensure timezone-naive datetime for consistent filename generation
        now = datetime.now()
        if hasattr(now, 'tzinfo') and now.tzinfo is not None:
            now = now.replace(tzinfo=None)
        timestamp = now.strftime("%Y%m%d_%H%M%S")
        
        if start_date or end_date:
            start_str = start_date.strftime("%Y%m%d") if start_date else "start"
            end_str = end_date.strftime("%Y%m%d") if end_date else "end"
            return f"expense_export_{start_str}_to_{end_str}_{timestamp}.xlsx"
        else:
            return f"expense_export_all_{timestamp}.xlsx"